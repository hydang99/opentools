# source code: https://github.com/inclusionAI/AWorld/blob/main/examples/gaia/mcp_collections/documents/msxlsx.py
import json, os, time, zipfile, sys, subprocess
from pathlib import Path
from typing import Any, Literal, Optional
from openpyxl import load_workbook
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__),'..', '..', '..')))
from opentools.core.base import BaseTool
from pydantic import Field, BaseModel
from pydantic.fields import FieldInfo
import pandas as pd


class DocumentMetadata(BaseModel): 
    """Metadata for extracted Excel document content."""
    file_name: str
    file_size: int
    file_type: str
    absolute_path: str
    page_count: Optional[int] = None
    processing_time: float
    extracted_images: list[str] = []
    extracted_media: list[dict[str, str]] = []
    output_format: str
    ocr_applied: bool
    extracted_text_file_path: Optional[str] = None


class Xlsxe_Extraction_Tool(BaseTool):
    """
    Xlsxe_Extraction_Tool
    ---------------------
    Purpose:
        A comprehensive Excel document extraction tool that extracts content from Excel files (.xlsx, .xls) with support for multiple worksheets, embedded media extraction, JPEG screenshot generation, and LLM-friendly output in multiple formats.

    Core Capabilities:
        - Excel content extraction from XLSX and XLS files using pandas with openpyxl/xlrd engines
        - Multiple worksheet support with selective sheet processing
        - Embedded media extraction (images, audio, video) from XLSX files
        - JPEG screenshot generation of Excel data for visual analysis

    Intended Use:
        Use this tool when you need to extract content from Excel files, including multiple worksheets, embedded media extraction, and JPEG screenshot generation.

    Limitations:
        - Image extraction only supported for XLSX files (not XLS)
        - Screenshot generation requires GUI environment with DISPLAY variable set (set create_screenshot=False for headless environments)
        - Large Excel files may consume significant memory
        - Screenshot quality depends on screen resolution
        - Media extraction limited to embedded content
        - Processing time scales with file size and complexity

    """
    # Default args for `opentools test Xlsxe_Extraction_Tool` (uses test_file/data.json)
    DEFAULT_TEST_ARGS = {
        "tool_test": "xls_extraction",
        "file_location": "xlsxe_extraction",
        "result_parameter": "result",
        "search_type": "similarity_eval",
    }

    def __init__(self) -> None:
        super().__init__(
            type='function',
            name="Xlsxe_Extraction_Tool",
            description="""A comprehensive Excel document extraction tool that extracts content from Excel files (.xlsx, .xls) with support for multiple worksheets, 
            embedded media extraction, JPEG screenshot generation, and LLM-friendly output in multiple formats. 
            CAPABILITIES: Excel content extraction from XLSX and XLS files using pandas with openpyxl/xlrd engines, 
            multiple worksheet support with selective sheet processing, embedded media extraction (images, audio, video) from XLSX files, 
            JPEG screenshot generation of Excel data for visual analysis. 
            SYNONYMS: Excel extraction tool, spreadsheet content extractor, XLSX data extractor, Excel worksheet processor, 
            spreadsheet media extractor, Excel screenshot generator, Excel data parser, spreadsheet analysis tool, 
            Excel content analyzer, worksheet extraction tool. 
            EXAMPLES: 'Extract content from spreadsheet.xlsx in markdown format', 'Extract specific sheets with embedded image extraction', 
            'Create screenshot of Excel data for visual analysis', 'Extract content from XLS file in JSON format including empty cells'.""",
            parameters={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel document file to extract content from"
                    },
                    "output_format": {
                        "type": "string",
                        "description": "Output format: 'markdown', 'json', 'html', or 'text' (default: json)",
                        "enum": ["markdown", "json", "html", "text"]
                    },
                    "extract_images": {
                        "type": "boolean",
                        "description": "Whether to extract and save embedded images from the document (default: true)"
                    },
                    "create_screenshot": {
                        "type": "boolean",
                        "description": "Whether to create a JPEG screenshot of the Excel data (default: true)"
                    },
                    "sheet_names": {
                        "type": ["string", "null"],
                        "description": "Comma-separated list of specific sheet names to process (null for all sheets)"
                    },
                    "include_empty_cells": {
                        "type": "boolean",
                        "description": "Whether to include empty cells in the output (default: false)"
                    },
                    "screenshot_max_rows": {
                        "type": "integer",
                        "description": "Maximum rows to include in screenshot (default: 50)"
                    },
                    "screenshot_max_cols": {
                        "type": "integer",
                        "description": "Maximum columns to include in screenshot (default: 20)"
                    }
                },
                "required": ["file_path"],
                "additionalProperties": False,
            },
            strict=False,
            category="document_extraction",
            tags=["excel_extraction", "spreadsheet_processing", "worksheet_extraction", "media_extraction", "screenshot_generation", "pandas", "openpyxl", "xlrd", "data_extraction", "document_processing"],
            limitation="Image extraction only supported for XLSX files (not XLS), screenshot generation requires GUI environment with DISPLAY variable set (set create_screenshot=False for headless environments), large Excel files may consume significant memory, screenshot quality depends on screen resolution, media extraction limited to embedded content, processing time scales with file size and complexity",
            agent_type="File_Extraction-Agent",
            accuracy= self.find_accuracy(os.path.join(os.path.dirname(__file__), 'test_result.json')),
            demo_commands= {
                "command": "reponse = tool.run(file_path='path/to/excel/file', output_format='markdown', extract_images=True, create_screenshot=True, sheet_names='Sheet1,Sheet2', include_empty_cells=True, screenshot_max_rows=50, screenshot_max_cols=20)",
                "description": "Extract content from excel file"
            }
        )
        self.workspace = Path(os.getcwd())
        self._media_output_dir = self.workspace / "extracted_media"

        # Create screenshots directory
        self._screenshots_dir = self.workspace / "excel_screenshots"

        self.supported_extensions: set = {
            ".xlsx",
            ".xls",
        }


    def _create_excel_screenshot(self, file_path: Path, sheet_name: str = None) -> str:
        """Create a JPEG screenshot of the valid Excel area using pyautogui.

        Args:
            file_path: Path to the Excel file
            sheet_name: Specific sheet to screenshot (None for first sheet)

        Returns:
            Path to the generated JPEG screenshot
        """
        try:
            # Lazy import pyautogui to avoid DISPLAY errors in headless environments
            try:
                import pyautogui
            except Exception as import_error:
                raise ValueError(
                    f"Screenshot functionality requires pyautogui and a DISPLAY environment. "
                    f"Error: {str(import_error)}. Set DISPLAY environment variable or disable create_screenshot."
                )
            
            self._screenshots_dir.mkdir(exist_ok=True)
            # Generate unique filename
            timestamp = int(time.time())
            screenshot_filename = f"{file_path.stem}_{sheet_name or 'sheet'}_{timestamp}.jpg"
            screenshot_path = self._screenshots_dir / screenshot_filename

            # Open Excel file with default application
            if sys.platform == "darwin":  # macOS
                subprocess.run(["open", str(file_path)], check=True)
            elif sys.platform == "win32":  # Windows
                subprocess.run(["start", str(file_path)], shell=True, check=True)
            else:  # Linux
                subprocess.run(["xdg-open", str(file_path)], check=True)

            # Wait for Excel to open
            time.sleep(3)

            # Take screenshot of the entire screen
            screenshot = pyautogui.screenshot()

            # Convert RGBA to RGB before saving as JPEG
            if screenshot.mode == "RGBA":
                screenshot = screenshot.convert("RGB")

            screenshot.save(screenshot_path, "JPEG", quality=95)

            return str(screenshot_path)

        except Exception as e:
            raise ValueError(f"Failed to create Excel screenshot with pyautogui: {str(e)}")
            

    def _extract_embedded_media_xlsx(self, file_path: Path) -> list[dict[str, str]]:
        """Extract embedded media from XLSX files.

        Args:
            file_path: Path to the XLSX file

        Returns:
            List of dictionaries containing media information
        """
        saved_media = []
        self._media_output_dir.mkdir(exist_ok=True)

        try:
            # Load workbook to extract images
            workbook = load_workbook(file_path, data_only=False)

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Extract images from worksheet
                if hasattr(worksheet, "_images"):
                    for idx, image in enumerate(worksheet._images):
                        try:
                            # Generate unique filename
                            image_filename = f"{file_path.stem}_{sheet_name}_img_{idx}.png"
                            image_path = self._media_output_dir / image_filename

                            # Save image
                            if hasattr(image, "ref"):
                                # Extract image data
                                img_data = image._data()
                                if img_data:
                                    with open(image_path, "wb") as f:
                                        f.write(img_data)

                                    saved_media.append(
                                        {
                                            "type": "image",
                                            "path": str(image_path),
                                            "sheet": sheet_name,
                                            "filename": image_filename,
                                        }
                                    )

                        except Exception as e:
                            raise ValueError(f"Failed to extract image {idx} from sheet {sheet_name}: {str(e)}")

            # Also try to extract from ZIP structure for additional media
            with zipfile.ZipFile(file_path, "r") as zip_file:
                media_files = [f for f in zip_file.namelist() if f.startswith("xl/media/")]

                for media_file in media_files:
                    try:
                        media_data = zip_file.read(media_file)
                        media_filename = f"{file_path.stem}_{Path(media_file).name}"
                        media_path = self._media_output_dir / media_filename

                        with open(media_path, "wb") as f:
                            f.write(media_data)

                        # Determine media type based on extension
                        media_ext = Path(media_file).suffix.lower()
                        if media_ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp"]:
                            media_type = "image"
                        elif media_ext in [".mp3", ".wav", ".m4a"]:
                            media_type = "audio"
                        elif media_ext in [".mp4", ".avi", ".mov"]:
                            media_type = "video"
                        else:
                            media_type = "other"

                        saved_media.append(
                            {
                                "type": media_type,
                                "path": str(media_path),
                                "filename": media_filename,
                                "original_path": media_file,
                            }
                        )

                    except Exception as e:
                        raise ValueError(f"Failed to extract media {media_file}: {str(e)}")

        except Exception as e:
            raise ValueError(f"Failed to extract media from XLSX: {str(e)}")

        return saved_media

    def _extract_excel_content(self, file_path: Path, sheet_names: list[str] | None = None) -> dict[str, Any]:
        """Extract content from Excel files using pandas and xlrd.

        Args:
            file_path: Path to the Excel file
            sheet_names: Specific sheets to process (None for all sheets)

        Returns:
            Dictionary containing extracted content and metadata
        """
        start_time = time.time()

        try:
            # Determine file type and read accordingly
            if file_path.suffix.lower() == ".xlsx":
                # Use openpyxl engine for XLSX files
                excel_file = pd.ExcelFile(file_path, engine="openpyxl")
            else:
                # Use xlrd engine for XLS files
                excel_file = pd.ExcelFile(file_path, engine="xlrd")

            # Get all sheet names if not specified
            if sheet_names is None:
                sheet_names = excel_file.sheet_names

            sheets_data = {}
            total_rows = 0
            total_cols = 0

            # Extract data from each sheet
            for sheet_name in sheet_names:
                if sheet_name in excel_file.sheet_names:
                    try:
                        # Read sheet data
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

                        # Remove completely empty rows and columns
                        df = df.dropna(how="all").dropna(axis=1, how="all")

                        if not df.empty:
                            sheets_data[sheet_name] = {
                                "data": df,
                                "shape": df.shape,
                                "columns": df.columns.tolist(),
                                "non_empty_cells": df.count().sum(),
                            }

                            total_rows += df.shape[0]
                            total_cols = max(total_cols, df.shape[1])
                        else:
                            sheets_data[sheet_name] = {"data": df, "shape": (0, 0), "columns": [], "non_empty_cells": 0}

                    except Exception as e:
                        print(f"Failed to read sheet '{sheet_name}': {str(e)}")
                        sheets_data[sheet_name] = {
                            "error": str(e),
                            "shape": (0, 0),
                            "columns": [],
                            "non_empty_cells": 0,
                        }

            processing_time = time.time() - start_time

            return {
                "sheets_data": sheets_data,
                "sheet_names": list(sheets_data.keys()),
                "total_sheets": len(sheets_data),
                "total_rows": total_rows,
                "total_columns": total_cols,
                "processing_time": processing_time,
                "file_engine": "openpyxl" if file_path.suffix.lower() == ".xlsx" else "xlrd",
            }

        except Exception as e:
            raise ValueError(f"Failed to extract Excel content: {str(e)}")

    def _format_content_for_llm(
        self, extraction_result: dict[str, Any], output_format: str, include_empty_cells: bool = False
    ) -> str:
        """Format extracted Excel content to be LLM-friendly.

        Args:
            extraction_result: Result from _extract_excel_content
            output_format: Desired output format
            include_empty_cells: Whether to include empty cells in output

        Returns:
            Formatted content string
        """
        sheets_data = extraction_result["sheets_data"]

        if output_format.lower() == "markdown":
            content_parts = []
            content_parts.append("# Excel Document Content\n")
            content_parts.append(f"**Total Sheets:** {extraction_result['total_sheets']}\n")
            content_parts.append(f"**Processing Engine:** {extraction_result['file_engine']}\n\n")

            for sheet_name, sheet_info in sheets_data.items():
                content_parts.append(f"## Sheet: {sheet_name}\n")

                if "error" in sheet_info:
                    content_parts.append(f"**Error:** {sheet_info['error']}\n\n")
                    continue

                df: pd.DataFrame = sheet_info["data"]
                shape = sheet_info["shape"]

                content_parts.append(f"**Dimensions:** {shape[0]} rows × {shape[1]} columns\n")
                content_parts.append(f"**Non-empty cells:** {sheet_info['non_empty_cells']}\n\n")

                if not df.empty:
                    # Convert DataFrame to markdown table
                    if include_empty_cells:
                        # Fill NaN values with empty string for display
                        df_display = df.fillna("")
                    else:
                        # Keep NaN values as they are
                        df_display = df

                    # Convert to markdown table
                    try:
                        markdown_table = df_display.to_markdown(index=False, tablefmt="pipe")
                        content_parts.append(f"### Data:\n{markdown_table}\n\n")
                    except Exception:
                        # Fallback to string representation
                        content_parts.append(f"### Data (text format):\n```\n{df_display.to_string()}\n```\n\n")
                else:
                    content_parts.append("*Sheet is empty*\n\n")

            return "".join(content_parts)

        elif output_format.lower() == "json":
            json_data = {
                "document_info": {
                    "total_sheets": extraction_result["total_sheets"],
                    "total_rows": extraction_result["total_rows"],
                    "total_columns": extraction_result["total_columns"],
                    "processing_engine": extraction_result["file_engine"],
                },
                "sheets": {},
            }

            for sheet_name, sheet_info in sheets_data.items():
                if "error" in sheet_info:
                    json_data["sheets"][sheet_name] = {"error": sheet_info["error"], "shape": sheet_info["shape"]}
                    continue

                df = sheet_info["data"]

                if not df.empty:
                    # Convert DataFrame to records
                    if include_empty_cells:
                        df_records = df.fillna("").to_dict("records")
                    else:
                        df_records = df.to_dict("records")

                    json_data["sheets"][sheet_name] = {
                        "shape": sheet_info["shape"],
                        "non_empty_cells": sheet_info["non_empty_cells"],
                        "data": df_records,
                    }
                else:
                    json_data["sheets"][sheet_name] = {"shape": sheet_info["shape"], "non_empty_cells": 0, "data": []}

            return json.dumps(json_data, indent=2, default=str)

        elif output_format.lower() == "html":
            html_parts = []
            html_parts.append("<html><body>")
            html_parts.append("<h1>Excel Document Content</h1>")
            html_parts.append(f"<p><strong>Total Sheets:</strong> {extraction_result['total_sheets']}</p>")
            html_parts.append(f"<p><strong>Processing Engine:</strong> {extraction_result['file_engine']}</p>")

            for sheet_name, sheet_info in sheets_data.items():
                html_parts.append(f"<h2>Sheet: {sheet_name}</h2>")

                if "error" in sheet_info:
                    html_parts.append(f"<p><strong>Error:</strong> {sheet_info['error']}</p>")
                    continue

                df = sheet_info["data"]
                shape = sheet_info["shape"]

                html_parts.append(f"<p><strong>Dimensions:</strong> {shape[0]} rows × {shape[1]} columns</p>")
                html_parts.append(f"<p><strong>Non-empty cells:</strong> {sheet_info['non_empty_cells']}</p>")

                if not df.empty:
                    # Convert DataFrame to HTML table
                    if include_empty_cells:
                        df_display = df.fillna("")
                    else:
                        df_display = df

                    html_table = df_display.to_html(index=False, escape=False, table_id=f"sheet_{sheet_name}")
                    html_parts.append(html_table)
                else:
                    html_parts.append("<p><em>Sheet is empty</em></p>")

            html_parts.append("</body></html>")
            return "".join(html_parts)

        else:  # text format
            content_parts = []
            content_parts.append(f"Excel Document Content\n{'=' * 50}\n")
            content_parts.append(f"Total Sheets: {extraction_result['total_sheets']}\n")
            content_parts.append(f"Processing Engine: {extraction_result['file_engine']}\n\n")

            for sheet_name, sheet_info in sheets_data.items():
                content_parts.append(f"Sheet: {sheet_name}\n{'-' * 30}\n")

                if "error" in sheet_info:
                    content_parts.append(f"Error: {sheet_info['error']}\n\n")
                    continue

                df = sheet_info["data"]
                shape = sheet_info["shape"]

                content_parts.append(f"Dimensions: {shape[0]} rows × {shape[1]} columns\n")
                content_parts.append(f"Non-empty cells: {sheet_info['non_empty_cells']}\n\n")

                if not df.empty:
                    if include_empty_cells:
                        df_display = df.fillna("")
                    else:
                        df_display = df

                    content_parts.append(f"Data:\n{df_display.to_string()}\n\n")
                else:
                    content_parts.append("Sheet is empty\n\n")

            return "".join(content_parts)

    def _validate_file_path(self, file_path: str) -> Path:
        """Validate and resolve file path. Rely on the predefined supported_extensions class variable.

        Args:
            file_path: Path to the document or media file

        Returns:
            Resolved Path object

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file type is not supported
        """
        path = Path(file_path).expanduser()
        if not path.is_absolute():
            path = self.workspace / path

        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        if path.suffix.lower() not in self.supported_extensions:
            raise ValueError(
                f"Unsupported file type: {path.suffix}. Supported types: {', '.join(self.supported_extensions)}"
            )

        return path
    
    def run(
        self,
        file_path: str = Field(description="Path to the Excel document file to extract content from"),
        output_format: Literal["markdown", "json", "html", "text"] = Field(
            default="json", description="Output format: 'markdown', 'json', 'html', or 'text'"
        ),
        extract_images: bool = Field(default=True, description="Whether to extract and save images from the document"),
        create_screenshot: bool = Field(
            default=True, description="Whether to create a JPEG screenshot of the Excel data"
        ),
        sheet_names: str | None = Field(
            default=None, description="Comma-separated list of specific sheet names to process (None for all sheets)"
        ),
        include_empty_cells: bool = Field(default=False, description="Whether to include empty cells in the output"),
        screenshot_max_rows: int = Field(default=50, description="Maximum rows to include in screenshot"),
        screenshot_max_cols: int = Field(default=20, description="Maximum columns to include in screenshot"),
    ):
        """Extract content from Excel documents using pandas and xlrd.

        This tool provides comprehensive Excel document content extraction with support for:
        - XLSX and XLS files
        - Multiple worksheets
        - Text and numeric data extraction
        - Image and media extraction (XLSX only)
        - JPEG screenshot generation of Excel data
        - Metadata collection
        - LLM-optimized output formatting

        Args:
            file_path: Path to the Excel file
            output_format: Desired output format
            extract_images: Whether to extract embedded images
            create_screenshot: Whether to create a JPEG screenshot
            sheet_names: Specific sheets to process
            include_empty_cells: Whether to include empty cells
            screenshot_max_rows: Maximum rows in screenshot
            screenshot_max_cols: Maximum columns in screenshot

        Returns:
            ActionResponse with extracted content, metadata, media file paths, and screenshot path
        """
        try:
            # Handle FieldInfo objects
            if isinstance(file_path, FieldInfo):
                file_path = file_path.default
            if isinstance(output_format, FieldInfo):
                output_format = output_format.default
            if isinstance(extract_images, FieldInfo):
                extract_images = extract_images.default
            if isinstance(create_screenshot, FieldInfo):
                create_screenshot = create_screenshot.default
            if isinstance(sheet_names, FieldInfo):
                sheet_names = sheet_names.default
            if isinstance(include_empty_cells, FieldInfo):
                include_empty_cells = include_empty_cells.default
            if isinstance(screenshot_max_rows, FieldInfo):
                screenshot_max_rows = screenshot_max_rows.default
            if isinstance(screenshot_max_cols, FieldInfo):
                screenshot_max_cols = screenshot_max_cols.default

            # Validate input file
            file_path: Path = self._validate_file_path(file_path)
            print(f"Processing Excel document: {file_path.name}")

            # Parse sheet names if provided
            target_sheets = None
            if sheet_names:
                target_sheets = [name.strip() for name in sheet_names.split(",")]

            # Extract content from Excel file
            extraction_result = self._extract_excel_content(file_path, target_sheets)

            # Extract embedded media if requested (XLSX only)
            saved_media = []
            if extract_images and file_path.suffix.lower() == ".xlsx":
                saved_media = self._extract_embedded_media_xlsx(file_path)
            elif extract_images and file_path.suffix.lower() == ".xls":
                print("Image extraction not supported for XLS files")

            # Create screenshot if requested
            screenshot_path = None
            if create_screenshot:
                try:
                    target_sheet = target_sheets[0] if target_sheets else None
                    screenshot_path = self._create_excel_screenshot(file_path, target_sheet)
                except Exception as e:
                    print(f"Warning: Could not create screenshot: {e}")
                    print("Continuing without screenshot...")
                    screenshot_path = None

            # Format content for LLM consumption
            formatted_content = self._format_content_for_llm(extraction_result, output_format, include_empty_cells)

            # Prepare metadata
            file_stats = file_path.stat()

            # Create Excel-specific metadata
            excel_metadata = {
                "sheet_count": extraction_result["total_sheets"],
                "sheet_names": extraction_result["sheet_names"],
                "total_rows": extraction_result["total_rows"],
                "total_columns": extraction_result["total_columns"],
                "processing_engine": extraction_result["file_engine"],
                "extracted_images": [media["path"] for media in saved_media if media["type"] == "image"],
                "extracted_media": saved_media,
                "screenshot_path": screenshot_path,
                "include_empty_cells": include_empty_cells,
                "processed_sheets": target_sheets or extraction_result["sheet_names"],
            }

            document_metadata = DocumentMetadata(
                file_name=file_path.name,
                file_size=file_stats.st_size,
                file_type=file_path.suffix.lower(),
                absolute_path=str(file_path.absolute()),
                page_count=extraction_result["total_sheets"],  # Use sheet count as "page" count
                processing_time=extraction_result["processing_time"],
                extracted_images=[media["path"] for media in saved_media if media["type"] == "image"],
                extracted_media=saved_media,
                output_format=output_format,
                llm_enhanced=False,
                ocr_applied=False,
            )

            # Combine standard and Excel-specific metadata
            combined_metadata = document_metadata.model_dump()
            combined_metadata.update(excel_metadata)

            success_message = (
                f"Successfully extracted content from {file_path.name} "
                f"({len(formatted_content)} characters, {extraction_result['total_sheets']} sheets, "
                f"{len(saved_media)} media files"
            )

            if screenshot_path:
                success_message += f", screenshot saved to: {screenshot_path}"


            return {'metadata': combined_metadata, 'result': formatted_content,  'success': True}

        except FileNotFoundError as e:
            return {'message': f"File not found: {e}", 'success': False}
        except ValueError as e:
            return {'message': f"Invalid input: {e}", 'success': False}
        except Exception as e:
             return {'message': f"Extraction error: {e}", 'success': False}
    
    def test(self, tool_test: str="xls_extraction", file_location: str="xlsxe_extraction", result_parameter: str="result", search_type: str="exact_match"):
        return super().test(tool_test=tool_test, file_location=file_location, result_parameter=result_parameter, search_type=search_type)

# Example usage and entry point
if __name__ == "__main__":

    # Initialize and run the Excel extraction service
    try:
        tool = Xlsxe_Extraction_Tool()
        tool.embed_tool()
        tool.test(tool_test="xls_extraction", file_location="xlsxe_extraction", result_parameter="result", search_type="exact_match")
    except Exception as e:
        print(f"Error: {e}")
